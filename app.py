"""
Aplicación Flask para reconocimiento OCR de matrículas y cuentakilómetros.
Utiliza OpenCV y Google Gemini Vision.
"""

from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file
import base64
import io
from PIL import Image
import cv2
import numpy as np
from ocr_processor import OCRProcessor
import os
from functools import wraps
from dotenv import load_dotenv
from datetime import datetime

# Cargar variables de entorno
load_dotenv()

# Inicializar Flask
app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # Límite de 16MB para imágenes
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', os.urandom(24))  # Clave secreta para sesiones

# Cargar usuarios desde .env o usar valores por defecto
def cargar_usuarios():
    """Carga usuarios desde variables de entorno"""
    usuarios_env = os.getenv('LOGIN_USERS', 'admin:admin123,user:user123')
    usuarios = {}
    for par in usuarios_env.split(','):
        if ':' in par:
            usuario, password = par.strip().split(':', 1)
            usuarios[usuario] = password
    return usuarios

USUARIOS = cargar_usuarios()

# Inicializar procesador OCR con Gemini
ocr_gemini = None

def get_ocr_processor():
    """Obtiene el procesador OCR de Gemini"""
    global ocr_gemini
    
    if ocr_gemini is None:
        app.logger.info("Inicializando OCR Processor...")
        try:
            ocr_gemini = OCRProcessor(motor='gemini')
            app.logger.info("OCR Processor inicializado correctamente")
        except Exception as e:
            app.logger.error(f"Error al inicializar OCR Processor: {str(e)}")
            raise
    return ocr_gemini


def login_required(f):
    """Decorador para requerir login en las rutas"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'username' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


@app.route('/login', methods=['GET', 'POST'])
def login():
    """Página de login"""
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if username in USUARIOS and USUARIOS[username] == password:
            session['username'] = username
            return redirect(url_for('index'))
        else:
            return render_template('login.html', error='Usuario o contraseña incorrectos')
    
    return render_template('login.html')


@app.route('/logout')
def logout():
    """Cerrar sesión"""
    session.pop('username', None)
    return redirect(url_for('login'))


@app.route('/')
@login_required
def index():
    """
    Redirige a la pantalla de gestión de vehículos.
    """
    return redirect(url_for('vehiculos'))


@app.route('/vehiculos')
@login_required
def vehiculos():
    """
    Pantalla de gestión de vehículos.
    Muestra tabla con matrículas y kilometrajes detectados.
    """
    # Inicializar lista de vehículos en sesión si no existe
    if 'vehiculos' not in session:
        session['vehiculos'] = []
    
    return render_template('vehiculos.html', 
                         username=session.get('username'),
                         vehiculos=session.get('vehiculos', []))


@app.route('/captura')
@login_required
def captura():
    """
    Página de captura de fotos (matrícula y kilometraje).
    """
    return render_template('captura.html', username=session.get('username'))


@app.route('/ocr/matricula', methods=['POST'])
@login_required
def procesar_matricula():
    """
    Endpoint para procesar imágenes de matrículas.
    
    Espera un JSON con la imagen en base64:
    {
        "image": "data:image/jpeg;base64,..."
    }
    
    Retorna JSON con el resultado:
    {
        "success": true,
        "tipo": "matricula",
        "texto": "1234ABC",
        "confianza": 85.5,
        "mensaje": "Matrícula reconocida correctamente"
    }
    """
    try:
        # Obtener datos de la imagen
        data = request.get_json()
        if not data or 'image' not in data:
            return jsonify({
                'success': False,
                'error': 'No se recibió ninguna imagen'
            }), 400
        
        # Decodificar imagen base64
        image_data = data['image']
        if image_data.startswith('data:image'):
            image_data = image_data.split(',')[1]
        
        image_bytes = base64.b64decode(image_data)
        image = Image.open(io.BytesIO(image_bytes))
        
        # Convertir a formato OpenCV
        img_cv = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)
        
        # Procesar matrícula con Gemini
        ocr = get_ocr_processor()
        resultado = ocr.procesar_matricula(img_cv)
        
        return jsonify(resultado)
    
    except Exception as e:
        app.logger.error(f"Error procesando matrícula: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error al procesar la imagen: {str(e)}'
        }), 500


@app.route('/ocr/cuentakilometros', methods=['POST'])
@login_required
def procesar_cuentakilometros():
    """
    Endpoint para procesar imágenes de cuentakilómetros.
    
    Espera un JSON con la imagen en base64:
    {
        "image": "data:image/jpeg;base64,..."
    }
    
    Retorna JSON con el resultado:
    {
        "success": true,
        "tipo": "cuentakilometros",
        "texto": "123456",
        "confianza": 90.2,
        "mensaje": "Lectura reconocida correctamente"
    }
    """
    try:
        app.logger.info("Recibida petición para procesar cuentakilómetros")
        
        # Obtener datos de la imagen
        data = request.get_json()
        if not data or 'image' not in data:
            app.logger.error("No se recibió imagen en la petición")
            return jsonify({
                'exito': False,
                'error': 'No se recibió ninguna imagen'
            }), 400
        
        # Decodificar imagen base64
        image_data = data['image']
        if image_data.startswith('data:image'):
            image_data = image_data.split(',')[1]
        
        image_bytes = base64.b64decode(image_data)
        image = Image.open(io.BytesIO(image_bytes))
        
        # Convertir a formato OpenCV
        img_cv = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)
        
        # Procesar cuentakilómetros con Gemini
        app.logger.info("Procesando con motor: Gemini")
        ocr = get_ocr_processor()
        resultado = ocr.procesar_cuentakilometros(img_cv)
        
        app.logger.info(f"Resultado OCR: {resultado}")
        
        return jsonify(resultado)
    
    except Exception as e:
        app.logger.error(f"Error procesando cuentakilómetros: {str(e)}", exc_info=True)
        return jsonify({
            'exito': False,
            'error': f'Error al procesar la imagen: {str(e)}'
        }), 500


@app.route('/health', methods=['GET'])
def health_check():
    """
    Endpoint para verificar el estado de la aplicación.
    """
    return jsonify({
        'status': 'ok',
        'message': 'Aplicación OCR funcionando correctamente'
    })


@app.route('/debug/config', methods=['GET'])
def debug_config():
    """
    Endpoint de debug para verificar configuración (solo para desarrollo).
    NO usar en producción sin autenticación adicional.
    """
    api_key = os.getenv('GEMINI_API_KEY', '')
    return jsonify({
        'gemini_api_key_configured': bool(api_key),
        'gemini_api_key_length': len(api_key) if api_key else 0,
        'gemini_api_key_preview': f"{api_key[:8]}...{api_key[-4:]}" if len(api_key) > 12 else 'TOO_SHORT',
        'secret_key_configured': bool(os.getenv('SECRET_KEY')),
        'login_users_configured': bool(os.getenv('LOGIN_USERS')),
        'port': os.getenv('PORT', '5002'),
        'environment_vars_count': len([k for k in os.environ.keys() if not k.startswith('_')])
    })


@app.route('/agregar_vehiculo', methods=['POST'])
@login_required
def agregar_vehiculo():
    """
    Agrega un vehículo a la lista en sesión.
    Espera JSON con matricula y kilometros.
    """
    try:
        data = request.get_json()
        matricula = data.get('matricula', '')
        kilometros = data.get('kilometros', '')
        
        if not matricula or not kilometros:
            return jsonify({
                'success': False,
                'error': 'Se requieren matrícula y kilómetros'
            }), 400
        
        # Inicializar lista si no existe
        if 'vehiculos' not in session:
            session['vehiculos'] = []
        
        # Agregar vehículo
        vehiculo = {
            'matricula': matricula,
            'kilometros': kilometros,
            'fecha': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        session['vehiculos'].append(vehiculo)
        session.modified = True
        
        return jsonify({
            'success': True,
            'vehiculo': vehiculo
        })
    
    except Exception as e:
        app.logger.error(f"Error agregando vehículo: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/editar_vehiculo', methods=['POST'])
@login_required
def editar_vehiculo():
    """
    Edita un vehículo existente en la lista.
    Espera JSON con index, matricula y kilometros.
    """
    try:
        data = request.get_json()
        index = data.get('index')
        matricula = data.get('matricula', '')
        kilometros = data.get('kilometros', '')
        
        if index is None or not matricula or not kilometros:
            return jsonify({
                'success': False,
                'error': 'Se requieren índice, matrícula y kilómetros'
            }), 400
        
        # Validar que existe la lista y el índice
        if 'vehiculos' not in session or index >= len(session['vehiculos']):
            return jsonify({
                'success': False,
                'error': 'Vehículo no encontrado'
            }), 404
        
        # Actualizar vehículo
        session['vehiculos'][index]['matricula'] = matricula
        session['vehiculos'][index]['kilometros'] = kilometros
        session.modified = True
        
        return jsonify({
            'success': True,
            'vehiculo': session['vehiculos'][index]
        })
    
    except Exception as e:
        app.logger.error(f"Error editando vehículo: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/eliminar_vehiculo', methods=['POST'])
@login_required
def eliminar_vehiculo():
    """
    Elimina un vehículo de la lista.
    Espera JSON con index.
    """
    try:
        data = request.get_json()
        index = data.get('index')
        
        if index is None:
            return jsonify({
                'success': False,
                'error': 'Se requiere el índice del vehículo'
            }), 400
        
        # Validar que existe la lista y el índice
        if 'vehiculos' not in session or index >= len(session['vehiculos']):
            return jsonify({
                'success': False,
                'error': 'Vehículo no encontrado'
            }), 404
        
        # Eliminar vehículo
        vehiculo_eliminado = session['vehiculos'].pop(index)
        session.modified = True
        
        return jsonify({
            'success': True,
            'vehiculo': vehiculo_eliminado
        })
    
    except Exception as e:
        app.logger.error(f"Error eliminando vehículo: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/descargar_excel')
@login_required
def descargar_excel():
    """
    Genera y descarga un archivo Excel con los vehículos registrados.
    """
    try:
        # Importar openpyxl solo cuando se necesita
        from openpyxl import Workbook
        
        vehiculos = session.get('vehiculos', [])
        
        if not vehiculos:
            return jsonify({
                'success': False,
                'error': 'No hay vehículos para descargar'
            }), 400
        
        # Crear libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Vehículos"
        
        # Encabezados
        ws['A1'] = 'Matrícula'
        ws['B1'] = 'Kilometraje'
        ws['C1'] = 'Fecha de Registro'
        
        # Aplicar formato a encabezados
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)
        
        # Agregar datos
        for idx, vehiculo in enumerate(vehiculos, start=2):
            ws[f'A{idx}'] = vehiculo.get('matricula', '')
            ws[f'B{idx}'] = vehiculo.get('kilometros', '')
            ws[f'C{idx}'] = vehiculo.get('fecha', '')
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Nombre del archivo con fecha
        filename = f'vehiculos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        app.logger.error(f"Error generando Excel: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


if __name__ == '__main__':
    # Ejecutar en modo debug para desarrollo
    # En producción, usar un servidor WSGI como Gunicorn
    import socket
    import os
    
    # Verificar si existen certificados SSL
    tiene_ssl = os.path.exists('cert.pem') and os.path.exists('key.pem')
    
    # Obtener la IP local
    try:
        hostname = socket.gethostname()
        local_ip = socket.gethostbyname(hostname)
    except:
        local_ip = "tu-ip-local"
    
    protocolo = "https" if tiene_ssl else "http"
    puerto = 5000
    
    print("=" * 60)
    print("🚀 Iniciando servidor Flask...")
    print("📷 Aplicación OCR - Matrículas y Cuentakilómetros")
    print("🔄 Motores OCR: Tesseract + OCR.space API + Gemini 2.5 Flash")
    print(f"🔒 Protocolo: {protocolo.upper()}")
    print("=" * 60)
    print()
    print("🌐 Accede desde este equipo:")
    print(f"   → {protocolo}://localhost:{puerto}")
    print(f"   → {protocolo}://127.0.0.1:{puerto}")
    print()
    print("📱 Accede desde otros dispositivos en la red local:")
    print(f"   → {protocolo}://{local_ip}:{puerto}")
    print()
    
    if tiene_ssl:
        print("✅ HTTPS ACTIVADO - Compatible con cámaras móviles")
        print("⚠️  IMPORTANTE:")
        print("   - El certificado es autofirmado")
        print("   - El navegador mostrará advertencia de seguridad")
        print("   - Acepta la advertencia para continuar")
        print("   - En móvil: 'Avanzado' → 'Continuar de todas formas'")
    else:
        print("⚠️  HTTP (sin cifrado)")
        print("   - Algunos navegadores móviles bloquearán la cámara")
        print("   - Para activar HTTPS ejecuta:")
        print("     python generar_certificado.py")
        print("   - Luego reinicia la aplicación")
    
    print()
    print("⚠️  Configuración del Firewall:")
    print("   - Asegúrate de que el firewall permita conexiones")
    print(f"     en el puerto {puerto}")
    print("   - Los dispositivos deben estar en la misma red WiFi")
    print("=" * 60)
    print()
    print("🔑 Credenciales de acceso:")
    print("   - Usuario: admin / Contraseña: admin123")
    print("   - Usuario: user / Contraseña: user123")
    print()
    
    # Configurar SSL si los certificados existen
    ssl_context = None
    if tiene_ssl:
        ssl_context = ('cert.pem', 'key.pem')
    
    # host='0.0.0.0' permite acceso desde cualquier IP en la red
    app.run(
        debug=True, 
        host='0.0.0.0', 
        port=puerto, 
        threaded=True,
        ssl_context=ssl_context
    )

# Para producción con gunicorn, no se ejecuta el bloque if __name__
# Gunicorn usa directamente la variable 'app'
